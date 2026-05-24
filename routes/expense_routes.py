import csv
import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, Response, send_file
from flask_login import login_required, current_user
from datetime import datetime, date
from models import db
from models.expense import Expense
from models.budget import Budget
from forms.expense_form import ExpenseForm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF
from sqlalchemy import func

expense_bp = Blueprint('expense', __name__)

@expense_bp.route('/', methods=['GET'])
@login_required
def list_expenses():
    """Paginated lists with sorting and searching filters."""
    # Read query parameters
    q_category = request.args.get('category', '').strip()
    q_search = request.args.get('search', '').strip()
    q_date_from = request.args.get('date_from', '').strip()
    q_date_to = request.args.get('date_to', '').strip()
    q_sort = request.args.get('sort', 'latest').strip()
    page = request.args.get('page', 1, type=int)
    
    query = Expense.query.filter_by(user_id=current_user.id)
    
    # 1. Filter by Category
    if q_category:
        query = query.filter(Expense.category == q_category)
        
    # 2. Search by Description
    if q_search:
        query = query.filter(Expense.description.ilike(f'%{q_search}%'))
        
    # 3. Filter by Date range
    if q_date_from:
        try:
            date_from = datetime.strptime(q_date_from, '%Y-%m-%d').date()
            query = query.filter(Expense.date >= date_from)
        except ValueError:
            pass
            
    if q_date_to:
        try:
            date_to = datetime.strptime(q_date_to, '%Y-%m-%d').date()
            query = query.filter(Expense.date <= date_to)
        except ValueError:
            pass
            
    # 4. Sorting logic
    if q_sort == 'oldest':
        query = query.order_by(Expense.date.asc(), Expense.id.asc())
    elif q_sort == 'highest':
        query = query.order_by(Expense.amount.desc(), Expense.id.desc())
    elif q_sort == 'lowest':
        query = query.order_by(Expense.amount.asc(), Expense.id.asc())
    else:  # 'latest' default
        query = query.order_by(Expense.date.desc(), Expense.id.desc())
        
    # Paginate results
    pagination = query.paginate(page=page, per_page=10, error_out=False)
    expenses = pagination.items
    
    # Quick Add Form
    form = ExpenseForm()
    
    # Predefined categories mapping
    categories_list = [
        ('Food', '🍔 Food'),
        ('Travel', '✈️ Travel'),
        ('Shopping', '🛍️ Shopping'),
        ('Bills', '💡 Bills'),
        ('Entertainment', '🎬 Entertainment'),
        ('Healthcare', '🏥 Healthcare'),
        ('Education', '📚 Education'),
        ('Others', '📦 Others')
    ]
    
    return render_template(
        'expenses.html',
        expenses=expenses,
        pagination=pagination,
        form=form,
        categories=categories_list,
        selected_category=q_category,
        search_query=q_search,
        date_from=q_date_from,
        date_to=q_date_to,
        sort_order=q_sort
    )


@expense_bp.route('/add', methods=['POST'])
@login_required
def add_expense():
    """CRUD: Add a new expense with support for custom transaction currencies."""
    form = ExpenseForm()
    if form.validate_on_submit():
        expense_date = form.date.data
        original_amount = float(form.amount.data)
        expense_currency = form.currency.data
        user_base_currency = current_user.currency
        
        # Real-time realistic exchange rates mapping (aligned with dashboard conversions)
        rates = {
            'USD': 1.0,
            'EUR': 0.92,
            'GBP': 0.79,
            'INR': 83.3,
            'JPY': 156.0,
            'AUD': 1.50,
            'CAD': 1.36,
            'CNY': 7.24
        }
        
        # Calculate amount converted to user's preferred base currency
        base_rate = rates.get(expense_currency, 1.0)
        amount_in_usd = original_amount / base_rate if base_rate > 0 else original_amount
        target_rate = rates.get(user_base_currency, 1.0)
        amount_in_base = amount_in_usd * target_rate
        
        expense = Expense(
            original_amount=original_amount,
            currency=expense_currency,
            amount=amount_in_base,
            category=form.category.data,
            description=form.description.data.strip() if form.description.data else '',
            date=expense_date,
            user_id=current_user.id
        )
        db.session.add(expense)
        db.session.commit()
        
        # Check Budget warning for category
        check_budget_warning(form.category.data, expense_date.month, expense_date.year)
        
        flash('Expense recorded successfully!', 'success')
    else:
        # Loop over form error messages
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text}: {error}", 'danger')
                
    return redirect(request.referrer or url_for('dashboard.index'))


@expense_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_expense(id):
    """CRUD: Edit existing expense and handles currency conversions cleanly."""
    expense = Expense.query.get_or_404(id)
    if expense.user_id != current_user.id:
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('expense.list_expenses'))
        
    form = ExpenseForm(obj=expense)
    if request.method == 'GET':
        form.amount.data = expense.original_amount
        form.currency.data = expense.currency
        
    if form.validate_on_submit():
        original_amount = float(form.amount.data)
        expense_currency = form.currency.data
        user_base_currency = current_user.currency
        
        rates = {
            'USD': 1.0,
            'EUR': 0.92,
            'GBP': 0.79,
            'INR': 83.3,
            'JPY': 156.0,
            'AUD': 1.50,
            'CAD': 1.36,
            'CNY': 7.24
        }
        
        # Calculate amount converted to user's preferred base currency
        base_rate = rates.get(expense_currency, 1.0)
        amount_in_usd = original_amount / base_rate if base_rate > 0 else original_amount
        target_rate = rates.get(user_base_currency, 1.0)
        amount_in_base = amount_in_usd * target_rate
        
        expense.original_amount = original_amount
        expense.currency = expense_currency
        expense.amount = amount_in_base
        expense.category = form.category.data
        expense.description = form.description.data.strip() if form.description.data else ''
        expense.date = form.date.data
        
        db.session.commit()
        
        # Check budget limit updates
        check_budget_warning(expense.category, expense.date.month, expense.date.year)
        
        flash('Expense updated successfully!', 'success')
        return redirect(url_for('expense.list_expenses'))
        
    return render_template('edit_expense.html', form=form, expense=expense)


@expense_bp.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_expense(id):
    """CRUD: Delete an expense."""
    expense = Expense.query.get_or_404(id)
    if expense.user_id != current_user.id:
        flash('Unauthorized action.', 'danger')
        return redirect(url_for('expense.list_expenses'))
        
    db.session.delete(expense)
    db.session.commit()
    flash('Expense has been deleted.', 'success')
    return redirect(url_for('expense.list_expenses'))


@expense_bp.route('/export/csv')
@login_required
def export_csv():
    """Generates and exports CSV records of all user expenses."""
    expenses = Expense.query.filter_by(user_id=current_user.id).order_by(Expense.date.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Date', 'Category', 'Description', 'Amount ($)', 'Created At'])
    
    for exp in expenses:
        writer.writerow([
            exp.id,
            exp.date.strftime('%Y-%m-%d'),
            exp.category,
            exp.description,
            exp.amount,
            exp.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
        
    response = Response(output.getvalue(), mimetype='text/csv')
    response.headers['Content-Disposition'] = 'attachment; filename=expenses_export.csv'
    return response


@expense_bp.route('/export/excel')
@login_required
def export_excel():
    """Generates premium Microsoft Excel formatted outputs."""
    expenses = Expense.query.filter_by(user_id=current_user.id).order_by(Expense.date.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses Summary"
    
    # Grid lines enabled
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = ['ID', 'Date', 'Category', 'Description', f'Amount ({current_user.currency})']
    ws.append(headers)
    
    # Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo brand
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    border_side = Side(border_style="thin", color="D1D5DB")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Style Header
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        
    # Append & Style Rows
    total_sum = 0.0
    for r_idx, exp in enumerate(expenses, start=2):
        ws.append([
            exp.id,
            exp.date.strftime('%Y-%m-%d'),
            exp.category,
            exp.description,
            exp.amount
        ])
        total_sum += exp.amount
        
        # Style cells
        ws.cell(row=r_idx, column=1).alignment = center_align
        ws.cell(row=r_idx, column=2).alignment = center_align
        ws.cell(row=r_idx, column=3).alignment = left_align
        ws.cell(row=r_idx, column=4).alignment = left_align
        
        amt_cell = ws.cell(row=r_idx, column=5)
        # Use user preferred currency symbol inside quotes in number format
        currency_symbol_safe = current_user.currency_symbol.replace('$', '\\$')
        amt_cell.number_format = f'"{currency_symbol_safe}"#,##0.00'
        amt_cell.alignment = right_align
        
        for c_idx in range(1, 6):
            ws.cell(row=r_idx, column=c_idx).border = thin_border
            ws.cell(row=r_idx, column=c_idx).font = Font(name="Segoe UI", size=10)
            
    # Add Total Sum row
    last_row = len(expenses) + 3
    ws.cell(row=last_row, column=4, value="Total Summary:").font = Font(name="Segoe UI", size=11, bold=True)
    ws.cell(row=last_row, column=4).alignment = right_align
    
    total_cell = ws.cell(row=last_row, column=5, value=total_sum)
    total_cell.font = Font(name="Segoe UI", size=11, bold=True, color="4F46E5")
    # Use user preferred currency symbol inside quotes in number format
    currency_symbol_safe = current_user.currency_symbol.replace('$', '\\$')
    total_cell.number_format = f'"{currency_symbol_safe}"#,##0.00'
    total_cell.alignment = right_align
    total_cell.border = Border(top=Side(style="thin", color="000000"), bottom=Side(style="double", color="000000"))
    
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    
    return send_file(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="Expenses_Report.xlsx"
    )


class PDFExpensesReport(FPDF):
    """Custom aesthetic PDF generator."""
    def header(self):
        self.set_fill_color(79, 70, 229) # Brand Indigo
        self.rect(0, 0, 210, 35, 'F')
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", 'B', 16)
        self.cell(0, 15, "EXPENSE TRACKER REPORT", ln=True, align='C')
        self.set_font("Helvetica", '', 10)
        self.cell(0, 5, f"Monthly Financial Log - Generated on {datetime.now().strftime('%Y-%m-%d')}", ln=True, align='C')
        self.ln(12)

    def footer(self):
        self.set_y(-15)
        self.set_font("Helvetica", 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()} | Created securely by Expense Tracker Manager", align='C')


@expense_bp.route('/export/pdf')
@login_required
def export_pdf():
    """Generates customized elegant PDF reports using fpdf2."""
    expenses = Expense.query.filter_by(user_id=current_user.id).order_by(Expense.date.desc()).all()
    
    pdf = PDFExpensesReport()
    pdf.add_page()
    pdf.set_font("Helvetica", '', 10)
    pdf.set_text_color(31, 41, 55) # Dark gray
    
    # Subheading info
    pdf.set_font("Helvetica", 'B', 12)
    pdf.cell(0, 8, f"Expense Ledger for: {current_user.username}", ln=True)
    pdf.set_font("Helvetica", '', 10)
    pdf.cell(0, 6, f"Email Account: {current_user.email}", ln=True)
    pdf.cell(0, 6, f"Total Transactions: {len(expenses)} records", ln=True)
    pdf.ln(5)
    
    # Table Header
    pdf.set_fill_color(243, 244, 246) # Light gray bg
    pdf.set_font("Helvetica", 'B', 10)
    # Total width = 190
    pdf.cell(15, 8, "ID", border=1, align='C', fill=True)
    pdf.cell(30, 8, "Date", border=1, align='C', fill=True)
    pdf.cell(40, 8, "Category", border=1, align='C', fill=True)
    pdf.cell(75, 8, "Description", border=1, align='L', fill=True)
    pdf.cell(30, 8, f"Amount ({current_user.currency})", border=1, align='R', fill=True)
    pdf.ln()
    
    # Table Content
    pdf.set_font("Helvetica", '', 9)
    total_amount = 0.0
    for exp in expenses:
        total_amount += exp.amount
        
        # Clamp length of description to fit nicely
        desc = exp.description if exp.description else ''
        if len(desc) > 35:
            desc = desc[:32] + "..."
            
        pdf.cell(15, 7, str(exp.id), border=1, align='C')
        pdf.cell(30, 7, exp.date.strftime('%Y-%m-%d'), border=1, align='C')
        pdf.cell(40, 7, exp.category, border=1, align='C')
        pdf.cell(75, 7, desc, border=1, align='L')
        pdf.cell(30, 7, f"{exp.amount:,.2f} {current_user.currency}", border=1, align='R')
        pdf.ln()
        
    # Total row
    pdf.set_font("Helvetica", 'B', 10)
    pdf.cell(160, 8, "Total Expenditure Summarized:", border=1, align='R')
    pdf.set_text_color(79, 70, 229)
    pdf.cell(30, 8, f"{total_amount:,.2f} {current_user.currency}", border=1, align='R')
    
    pdf_output = pdf.output()
    # fpdf2 output() returns a bytearray if no file is specified. Wrap in bytes() for Werkzeug/Flask
    if isinstance(pdf_output, str):
        pdf_output = pdf_output.encode('latin1')
    pdf_stream = io.BytesIO(pdf_output)
    return send_file(pdf_stream, mimetype="application/pdf", as_attachment=True, download_name="expenses_report.pdf")


@expense_bp.route('/import/csv', methods=['POST'])
@login_required
def import_csv():
    """Allows CSV data entry validation imports."""
    file = request.files.get('file')
    if not file or not file.filename.endswith('.csv'):
        flash('Please upload a valid CSV file.', 'danger')
        return redirect(url_for('expense.list_expenses'))
        
    try:
        stream = io.StringIO(file.stream.read().decode("utf-8"), newline=None)
        csv_reader = csv.reader(stream)
        
        header = next(csv_reader, None)  # Skip header
        # Support basic header matching
        imported_count = 0
        
        for row in csv_reader:
            if not row or len(row) < 3:
                continue
                
            # Expecting Date, Category, Description, Amount
            # Or Date, Category, Amount, Description
            # Let's map dynamically or guess based on numeric columns
            try:
                date_str = row[0].strip()
                category = row[1].strip()
                
                # Check options
                if len(row) >= 4:
                    desc = row[2].strip()
                    amount = float(row[3].replace('$', '').replace(',', '').strip())
                else:
                    desc = ''
                    amount = float(row[2].replace('$', '').replace(',', '').strip())
                    
                # Format checks
                expense_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                
                # Clean categories match
                valid_categories = {'Food', 'Travel', 'Shopping', 'Bills', 'Entertainment', 'Healthcare', 'Education', 'Others'}
                if category not in valid_categories:
                    category = 'Others'
                    
                expense = Expense(
                    amount=amount,
                    category=category,
                    description=desc,
                    date=expense_date,
                    user_id=current_user.id
                )
                db.session.add(expense)
                imported_count += 1
            except Exception:
                continue
                
        db.session.commit()
        flash(f'Successfully imported {imported_count} expense entries!', 'success')
    except Exception as e:
        flash(f'Failed to import CSV: {str(e)}', 'danger')
        
    return redirect(url_for('expense.list_expenses'))


def check_budget_warning(category, month, year):
    """Triggers visual overspending warnings inside flash context."""
    budget = Budget.query.filter_by(user_id=current_user.id, category=category, month=month, year=year).first()
    if budget:
        # Sum of expense
        total_spend = Expense.query.filter(
            Expense.user_id == current_user.id,
            Expense.category == category,
            func.strftime('%m', Expense.date) == f'{month:02d}',
            func.strftime('%Y', Expense.date) == str(year)
        ).with_entities(func.sum(Expense.amount)).scalar() or 0.0
        
        if total_spend > budget.monthly_limit:
            flash(
                f"🚨 Overspending Alert! Your '{category}' expenses (${total_spend:,.2f}) "
                f"have exceeded your monthly limit of ${budget.monthly_limit:,.2f}!",
                "warning"
            )
